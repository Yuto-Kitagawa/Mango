import openpyxl as openpyxl
import xlrd
import numpy as np
import pandas as pd
import datetime


class Functions:
    def toNum(self, min_row, min_col, max_row, max_col):
        wb = openpyxl.load_workbook("./必要データ纏め.xlsx")
        sheet = wb["マスタデータレース一覧"]
        tenkou = []

        # 行単位でセルの値を取得
        for cols in sheet.iter_cols(min_col, max_col, min_row, max_row):
            for cell in cols:
                if cell.value == "晴":  # 晴れなら1,#曇りなら2,#その他3
                    sheet.cell(row=4, column=cols).value = 1
                    tenkou.append(1)
                elif cell.value == "雲":
                    tenkou.append(2)
                    sheet.cell(row=4, column=cols).value = 2
                else:
                    tenkou.append(3)
                    sheet.cell(row=4, column=cols).value = 3

        return tenkou

    def split_data(self, df):
        # df = pd.read_excel("./必要データ纏め.xlsx", sheet_name="RACEORDER")
        sorted_id_list = df.sort_values(by='連番').index.unique()
        train_id_list = sorted_id_list[:round(len(sorted_id_list)*0.7)]
        test_id_list = sorted_id_list[round(len(sorted_id_list)*0.7):]
        train = df.loc[train_id_list]
        test = df.loc[test_id_list]
        return train, test

    def toArray():
        # https://www.youtube.com/watch?v=LhCBY_3kZkw
        wb = xlrd.load_workbook("./hy_data.xlsx")
        sheet = wb["Sheet1"]
        Sheet_Max = Sheet_1.nrows

        # 配列宣言
        id = []
        frame = []
        horse_number = []
        kinryou = []
        course_len = []
        weather = []
        race_type = []
        course_status = []
        date = []
        gender = []
        age = []
        weight_change = []
        rank = []

        # シートの2行目~最終行をループ
        for i in range(1, Sheet_Max):
            id.append(int(Sheet_1.cell_value(i, 0)))
            frame.append(int(Sheet_1.cell_value(i, 1)))
            horse_number.append(int(Sheet_1.cell_value(i, 1)))
            kinryou.append(int(Sheet_1.cell_value(i, 1)))
            course_len.append(int(Sheet_1.cell_value(i, 1)))
            weather.append(Sheet_1.cell_value(i, 1))#文字
            race_type.append(Sheet_1.cell_value(i, 1))  # 文字
            weather.append(int(Sheet_1.cell_value(i, 1)))  # 文字
            course_status.append(int(Sheet_1.cell_value(i, 1)))  # 文字
            date.append(int(Sheet_1.cell_value(i, 1)))  # 文字型
            gender.append(int(Sheet_1.cell_value(i, 1)))  # 文字
            age.append(int(Sheet_1.cell_value(i, 1)))
            weight_change.append(int(Sheet_1.cell_value(i, 1)))
            rank.append(int(Sheet_1.cell_value(i, 1)))

        print(My_Name)
        print(My_Suryo)

        pass
